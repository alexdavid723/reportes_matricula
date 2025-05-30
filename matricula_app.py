import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.page import PageMargins
from datetime import date
from datetime import datetime
import os
import pythoncom
import win32com.client as win32

# Interfaz con Tkinter
root = tk.Tk()
root.title("App de Matrícula")
root.geometry("400x250")

nivel_seleccionado = tk.StringVar()

def procesar_archivo():
    try:
        file_path = filedialog.askopenfilename(title="Selecciona un archivo Excel", filetypes=[("Excel files", "*.xlsx")])
        if not file_path:
            return
        

        nivel = nivel_seleccionado.get()
        if nivel == "Inicial":
            columnas_a_eliminar = ["DRE", "UGEL", "Departamento", "Provincia","Nivel","DNI Validado","DNI sin Validar","Registrado sin DNI", "Distrito", "Centro Poblado",
                                   "0 años", "1 año", "2 años", "3 años", "4 años", "5 años", "más de 5 años", "Unnamed: 13", "Unnamed: 15", "Unnamed: 16", "Unnamed: 18", "Unnamed: 28", "Unnamed: 30"
                                   , "Unnamed: 32", "Unnamed: 34", "Unnamed: 36", "Unnamed: 38", "Unnamed: 40"]
           
            df = pd.read_excel(file_path, skiprows=7)
            df_filtrado = df[df["Matricula En Proceso"].fillna(0) != 0]

            df_filtrado = df_filtrado.reset_index(drop=True)
            df_filtrado.index = df_filtrado.index + 1
            df_filtrado = df_filtrado.rename_axis('No').reset_index()
            df_filtrado = df_filtrado.drop(columns=[col for col in columnas_a_eliminar if col in df_filtrado.columns])
        elif nivel == "Primaria":
            columnas_a_eliminar = ["DRE", "UGEL", "Departamento", "Provincia","Nivel","DNI Validado","DNI sin Validar","Registrado sin DNI", "Distrito", "Centro Poblado",
                                   "Primero", "Segundo", "Tercero", "Cuarto", "Quinto", "Sexto", "Unnamed: 17", "Unnamed: 13", "Unnamed: 15", "Unnamed: 16", "Unnamed: 18", "Unnamed: 28", "Unnamed: 30"
                                   , "Unnamed: 32", "Unnamed: 34", "Unnamed: 36", "Unnamed: 38", "Unnamed: 40"]
            
            df = pd.read_excel(file_path, skiprows=7)
            df_filtrado = df[df["Matricula En Proceso"].fillna(0) != 0]

            df_filtrado = df_filtrado.reset_index(drop=True)
            df_filtrado.index = df_filtrado.index + 1
            df_filtrado = df_filtrado.rename_axis('No').reset_index()
            df_filtrado = df_filtrado.drop(columns=[col for col in columnas_a_eliminar if col in df_filtrado.columns])
        elif nivel == "Secundaria":
            columnas_a_eliminar = ["DRE", "UGEL", "Departamento", "Provincia","Nivel","DNI Validado","DNI sin Validar","Registrado sin DNI", "Distrito", "Centro Poblado",
                                   "Primero", "Segundo", "Tercero", "Cuarto", "Quinto", "Unnamed: 17", "Unnamed: 13", "Unnamed: 15", "Unnamed: 16", "Unnamed: 18", "Unnamed: 28", "Unnamed: 30"
                                   , "Unnamed: 32", "Unnamed: 34", "Unnamed: 36", "Unnamed: 38", "Unnamed: 40"]
            
            df = pd.read_excel(file_path, skiprows=6)
            df_filtrado = df[df["Matricula En Proceso"].fillna(0) != 0]

            df_filtrado = df_filtrado.reset_index(drop=True)
            df_filtrado.index = df_filtrado.index + 1
            df_filtrado = df_filtrado.rename_axis('No').reset_index()
            df_filtrado = df_filtrado.drop(columns=[col for col in columnas_a_eliminar if col in df_filtrado.columns])
        wb = Workbook()
        ws = wb.active

        today = date.today().strftime("%Y-%m-%d")
        nivel_formateado = nivel.lower()
        titulo = f"Reporte de matrícula {today} - {nivel}"

        ws.insert_rows(1, 1)
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=len(df_filtrado.columns))
    
        for r in dataframe_to_rows(df_filtrado, index=False, header=True):
            ws.append(r)

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                cell.font = Font(size=8)
                cell.border = Border(left=Side(style='thin'),
                                     right=Side(style='thin'),
                                     top=Side(style='thin'),
                                     bottom=Side(style='thin'))

        # Encabezado verde claro
        green_fill = PatternFill(start_color="006400", end_color="006400", fill_type="solid")
        for cell in ws[3]:
            cell.fill = green_fill

        # Insertar fila vacía justo después del encabezado
        # Insertar fila vacía justo después del encabezado
        ws.insert_rows(4, 1)

        # Obtener el número de columnas totales
        num_columnas = len(df_filtrado.columns)

        # Determinar cuántas columnas serán agrupadas al final (últimas 3)
        columnas_total = len(df_filtrado.columns)
        col_inicio_nominas = columnas_total - 2  # Últimas 3 columnas

        # Estilos
        green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        header_font = Font(size=8, bold=True)
        regular_font = Font(size=8)

        # Recorrer todas las columnas
        for col in range(1, columnas_total + 1):
            cell = ws.cell(row=3, column=col)
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            cell.fill = green_fill
            cell.font = header_font

            # Para las columnas antes de "Nóminas de Matrícula", combinar verticalmente
            if col < col_inicio_nominas:
                ws.merge_cells(start_row=3, start_column=col, end_row=4, end_column=col)

        # Combinar las últimas 3 columnas horizontalmente para "Nóminas de Matrícula"
        ws.merge_cells(start_row=3, start_column=col_inicio_nominas, end_row=3, end_column=columnas_total)
        titulo_nominas = ws.cell(row=3, column=col_inicio_nominas)
        titulo_nominas.value = "Nóminas de Matrícula"
        titulo_nominas.alignment = Alignment(horizontal='center', vertical='center')
        titulo_nominas.fill = green_fill
        titulo_nominas.font = header_font

        # Escribir los títulos individuales debajo (fila 4)
        ws.cell(row=4, column=col_inicio_nominas).value = "Generadas"
        ws.cell(row=4, column=col_inicio_nominas + 1).value = "Aprobadas"
        ws.cell(row=4, column=col_inicio_nominas + 2).value = "Por Rectificar"

        for col in range(col_inicio_nominas, columnas_total + 1):
            cell = ws.cell(row=4, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.fill = green_fill
            cell.font = regular_font
            cell.border = Border(left=Side(style='thin'),
                                     right=Side(style='thin'),
                                     top=Side(style='thin'),
                                     bottom=Side(style='thin'))



        if "Matricula En Proceso" in df_filtrado.columns:
            idx = df_filtrado.columns.get_loc("Matricula En Proceso") + 1
            yellow_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
            for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=idx, max_col=idx):
                for cell in row:
                    cell.fill = yellow_fill 

        #ws.page_margins = PageMargins(left=1.78, right=1.78, top=1.91, bottom=1.91, header=0.76, footer=0.76)
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4

        title_cell = ws.cell(row=1, column=1, value=titulo)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.font = Font(size=18, bold=False)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        timestamp2 = datetime.now().strftime("%Y""-""%m""-""%d")
        nivel_formateado = nivel.lower()  # convierte "Inicial" a "inicial", etc.
        nombre_archivo = f"{timestamp2}_{nivel_formateado}.xlsx"
        output_path = os.path.join(os.path.dirname(file_path), nombre_archivo)
        wb.save(output_path)
                # Crear versión PDF desde Excel
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        wb_com = excel.Workbooks.Open(output_path)
        pdf_output_path = output_path.replace(".xlsx", ".pdf")
        ws_com = wb_com.Worksheets[1]
        ws_com.PageSetup.Zoom = False
        ws_com.PageSetup.FitToPagesTall = False
        ws_com.PageSetup.FitToPagesWide = 1
        wb_com.ExportAsFixedFormat(0, pdf_output_path)
        wb_com.Close(False)
        excel.Quit()

        messagebox.showinfo("Éxito", f"Archivo procesado y guardado como:\n{output_path}")
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    except Exception as e:
        messagebox.showerror("Error", str(e))

label = tk.Label(root, text="Procesar archivo de matrícula Excel", font=("Arial", 12))
label.pack(pady=10)

nivel_label = tk.Label(root, text="Selecciona el nivel educativo:", font=("Arial", 10))
nivel_label.pack()
combo = ttk.Combobox(root, textvariable=nivel_seleccionado, state="readonly", font=("Arial", 10))
combo['values'] = ["Inicial", "Primaria", "Secundaria"]
combo.current(0)
combo.pack(pady=5)

btn = tk.Button(root, text="Seleccionar y Procesar Archivo", command=procesar_archivo,
                font=("Arial", 11), bg="#4D964F", fg="white")
btn.pack(pady=15)

root.mainloop()
